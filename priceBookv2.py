# price_book.py
import os
import pandas as pd
import tldextract
from datetime import datetime
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class PriceBook:
    DEFAULT_XLSX = "price_of_fertilizers.xlsx"
    DATE_FMT = "%Y-%m-%d %H:%M:%S"

    BASE_COLUMNS = [
        "url", "name", "category",
        "price", "price_per_kg",
        "amount_kg", "last_price_update", "is_available",
        "created_at", "last_updated",
    ]

    def __init__(self, path: str | None = None):
        self.path = path or self.DEFAULT_XLSX
        self.sheets: dict[str, pd.DataFrame] = {}

        if os.path.exists(self.path):
            self.sheets = pd.read_excel(
                self.path,
                sheet_name=None,
                engine="openpyxl",
                dtype=str
            )
            for k, df in self.sheets.items():
                self.sheets[k] = df.reindex(columns=self.BASE_COLUMNS)

    def upsert(self, row: Dict[str, Any]) -> None:
        row = {k: str(v) for k, v in row.items()}
        domain = self._domain_from_url(row["url"])

        if domain not in self.sheets:
            self.sheets[domain] = pd.DataFrame(columns=self.BASE_COLUMNS)

        df = self.sheets[domain]
        pk = row["url"]
        now = datetime.now().strftime(self.DATE_FMT)
        mask = df["url"] == pk

        if mask.any():
            idx = df.index[mask][0]
            for k, v in row.items():
                df.at[idx, k] = v
            df.at[idx, "last_updated"] = now
        else:
            row |= {
                "created_at": now,
                "last_updated": now,
                "amount_kg": row.get("amount_kg", ""),
                "last_price_update": row.get("last_price_update", ""),
                "is_available": row.get("is_available", "")
            }
            self.sheets[domain].loc[len(df)] = row

        self.save()

    def bulk_upsert(self, rows: List[Dict[str, Any]]):
        for r in rows:
            self.upsert(r)

    def save(self, path: str | None = None) -> None:
        from openpyxl import load_workbook

        outfile = path or self.path
        with pd.ExcelWriter(outfile, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in self.sheets.items():
                df["price"] = df["price"].astype(str)
                df["price_per_kg"] = df["price_per_kg"].astype(str)
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        # Apply styling
        wb = load_workbook(outfile)

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        alt_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet in wb.worksheets:
            max_row = sheet.max_row

            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
                adjusted_width = max(14, min(50, max_length + 2))
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = adjusted_width

            # Header styling
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="000000")
                cell.alignment = Alignment(horizontal="center")
                cell.fill = header_fill
                cell.border = border

            # Data row styling
            for row in sheet.iter_rows(min_row=2, max_row=max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left")
                    cell.border = border
                if row[0].row % 2 == 0:
                    for cell in row:
                        cell.fill = alt_fill

        wb.save(outfile)
        print(f"✅ Saved {sum(len(df) for df in self.sheets.values())} rows → {outfile}")

    @staticmethod
    def _domain_from_url(url: str) -> str:
        ext = tldextract.extract(url)
        return ".".join(part for part in [ext.domain, ext.suffix] if part)
