# price_book.py
import os
import pandas as pd
import tldextract
from datetime import datetime
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class PriceBook:
    DEFAULT_XLSX = "price_of_fertilizers.xlsx"
    DATE_FMT = "%Y-%m-%d %H:%M:%S"

    BASE_COLUMNS = [
        "url", "name", "category",
        "price", "price_per_kg",
        "amount_kg", "last_price_update", "is_available",
        "created_at", "last_updated",
    ]

    def __init__(self, path: str | None = None):
        self.path = path or self.DEFAULT_XLSX
        self.sheets: dict[str, pd.DataFrame] = {}

        if os.path.exists(self.path):
            self.sheets = pd.read_excel(
                self.path,
                sheet_name=None,
                engine="openpyxl",
                dtype=str
            )
            for k, df in self.sheets.items():
                self.sheets[k] = df.reindex(columns=self.BASE_COLUMNS)

    def upsert(self, row: Dict[str, Any]) -> None:
        row = {k: str(v) for k, v in row.items()}
        domain = self._domain_from_url(row["url"])

        if domain not in self.sheets:
            self.sheets[domain] = pd.DataFrame(columns=self.BASE_COLUMNS)

        df = self.sheets[domain]
        pk = row["url"]
        now = datetime.now().strftime(self.DATE_FMT)
        mask = df["url"] == pk

        if mask.any():
            idx = df.index[mask][0]
            for k, v in row.items():
                df.at[idx, k] = v
            df.at[idx, "last_updated"] = now
        else:
            row |= {
                "created_at": now,
                "last_updated": now,
                "amount_kg": row.get("amount_kg", ""),
                "last_price_update": row.get("last_price_update", ""),
                "is_available": row.get("is_available", "")
            }
            self.sheets[domain].loc[len(df)] = row

        self.save()

    def bulk_upsert(self, rows: List[Dict[str, Any]]):
        for r in rows:
            self.upsert(r)

    def save(self, path: str | None = None) -> None:
        from openpyxl import load_workbook

        outfile = path or self.path
        with pd.ExcelWriter(outfile, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in self.sheets.items():
                df["price"] = df["price"].astype(str)
                df["price_per_kg"] = df["price_per_kg"].astype(str)
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        # Apply modern Excel styles
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = load_workbook(outfile)

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        normal_font = Font(name="Calibri", size=10)
        border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )

        for sheet in wb.worksheets:
            max_row = sheet.max_row
            max_col = sheet.max_column

            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(sheet.cell(row=r, column=col_idx).value or "")) for r in range(1, max_row + 1))
                sheet.column_dimensions[col_letter].width = max(14, min(50, max_len + 2))

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row in sheet.iter_rows(min_row=2, max_row=max_row):
                is_even = row[0].row % 2 == 0
                for cell in row:
                    cell.font = normal_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.fill = alt_fill if is_even else PatternFill(fill_type=None)
                    cell.border = border

            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

        wb.save(outfile)
        print(f"✅ Modernized and saved {sum(len(df) for df in self.sheets.values())} rows → {outfile}")


    @staticmethod
    def _domain_from_url(url: str) -> str:
        ext = tldextract.extract(url)
        return ".".join(part for part in [ext.domain, ext.suffix] if part)
