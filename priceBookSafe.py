# price_book.py
import os, re
import pandas as pd
import tldextract
from datetime import datetime
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

class PriceBook:
    """
    Maintain an Excel workbook of fertilizer prices.
    Each website/domain gets its own sheet.
    Primary key per sheet = 'url'.
    """

    DEFAULT_XLSX = "price_of_fertilizers.xlsx"
    DATE_FMT = "%Y-%m-%d %H:%M:%S"

    BASE_COLUMNS = [
        "url", "name", "category",
        "price", "price_per_kg",
        "amount_kg", "last_price_update", "is_available",
        "created_at", "last_updated",
    ]

    def __init__(self, path: str | None = None):
        self.path = path or self.DEFAULT_XLSX
        self.sheets: dict[str, pd.DataFrame] = {}

        if os.path.exists(self.path):
            self.sheets = pd.read_excel(
                self.path,
                sheet_name=None,
                engine="openpyxl",
                dtype=str
            )
            for k, df in self.sheets.items():
                self.sheets[k] = df.reindex(columns=self.BASE_COLUMNS)

    def upsert(self, row: Dict[str, Any]) -> None:
        row = {k: str(v) for k, v in row.items()}

        domain = self._domain_from_url(row["url"])
        if domain not in self.sheets:
            self.sheets[domain] = pd.DataFrame(columns=self.BASE_COLUMNS)

        df = self.sheets[domain]
        pk = row["url"]
        now = datetime.now().strftime(self.DATE_FMT)

        mask = df["url"] == pk
        if mask.any():
            idx = df.index[mask][0]
            for k, v in row.items():
                df.at[idx, k] = v
            df.at[idx, "last_updated"] = now
        else:
            row |= {
                "created_at": now,
                "last_updated": now,
                "amount_kg": row.get("amount_kg", ""),
                "last_price_update": row.get("last_price_update", ""),
                "is_available": row.get("is_available", "")
            }
            self.sheets[domain].loc[len(df)] = row

        self.save()

    def bulk_upsert(self, rows: List[Dict[str, Any]]):
        for r in rows:
            self.upsert(r)

    def save(self, path: str | None = None) -> None:
        from openpyxl import load_workbook

        outfile = path or self.path
        with pd.ExcelWriter(outfile, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in self.sheets.items():
                df["price"] = df["price"].astype(str)
                df["price_per_kg"] = df["price_per_kg"].astype(str)
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        # Resize columns and add styles
        from openpyxl import load_workbook
        wb = load_workbook(outfile)
        for sheet in wb.worksheets:
            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
                adjusted_width = max(12, min(40, max_length + 2))
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = adjusted_width

            # Header style
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        wb.save(outfile)
        print(f"✅ Saved {sum(len(df) for df in self.sheets.values())} rows → {outfile}")

    @staticmethod
    def _domain_from_url(url: str) -> str:
        ext = tldextract.extract(url)
        return ".".join(part for part in [ext.domain, ext.suffix] if part)
