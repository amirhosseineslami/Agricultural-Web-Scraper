# price_book.py
import os
import pandas as pd
import tldextract, traceback
from datetime import datetime
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class PriceBook:
    DEFAULT_XLSX = "price_of_fertilizers.xlsx"
    DEFAULT_XLSX = os.path.join("output", DEFAULT_XLSX)
    DATE_FMT = "%Y-%m-%d %H:%M:%S"

    BASE_COLUMNS = [
        "url",
        "name",
        "category",
        "price",
        "price_per_kg",
        "amount_kg",
        "last_price_update",
        "is_available",
        "created_at",
        "last_updated",
        "product_menu_url",
    ]

    def __init__(self, path: str | None = None):
        self.path = path or self.DEFAULT_XLSX
        self.sheets: dict[str, pd.DataFrame] = {}

        if os.path.exists(self.path):
            self.sheets = pd.read_excel(
                self.path, sheet_name=None, engine="openpyxl", dtype=str
            )
            for k, df in self.sheets.items():
                self.sheets[k] = df.reindex(columns=self.BASE_COLUMNS)

    def upsert(self, row: Dict[str, Any]) -> None:
        row = {k: str(v) for k, v in row.items()}
        domain = self._domain_from_url(row["url"])
        if len(domain) < 2:
            assert Exception("Domain is not available!")
            return

        if domain not in self.sheets:
            self.sheets[domain] = pd.DataFrame(columns=self.BASE_COLUMNS)

        df = self.sheets[domain]

        # 👇 Use a composite key for uniqueness
        pk = f"{row['url']}::{row['name']}"
        now = datetime.now().strftime(self.DATE_FMT)

        # Add new composite key column if not already present
        if "_pk" not in df.columns:
            df["_pk"] = df["url"].astype(str) + "::" + df["name"].astype(str)

        mask = df["_pk"] == pk

        if mask.any():
            idx = df.index[mask][0]
            for k, v in row.items():
                df.at[idx, k] = v
            df.at[idx, "last_updated"] = now
        else:
            row |= {
                "created_at": now,
                "last_updated": now,
                "amount_kg": row.get("amount_kg", ""),
                "last_price_update": row.get("last_price_update", ""),
                "is_available": row.get("is_available", ""),
                "_pk": pk,  # 👈 Add the composite key
            }
            self.sheets[domain].loc[len(df)] = row

        self.save()

    def bulk_upsert(self, rows: List[Dict[str, Any]]):
        for r in rows:
            self.upsert(r)

    @staticmethod
    def _domain_from_url(url: str) -> str:
        ext = tldextract.extract(url)
        return ".".join(part for part in [ext.domain, ext.suffix] if part)

    def save(self, path: str | None = None) -> None:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        outfile = path or self.path

        # Step 1: Write data to Excel
        with pd.ExcelWriter(outfile, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in self.sheets.items():
                df["price"] = df["price"].astype(str)
                df["price_per_kg"] = df["price_per_kg"].astype(str)
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        # Step 2: Load workbook to style
        wb = load_workbook(outfile)

        # Styling objects
        header_fill = PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        )
        alt_fill = PatternFill(
            start_color="F7F7F7", end_color="F7F7F7", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for sheet in wb.worksheets:
            max_row = sheet.max_row
            max_col = sheet.max_column

            # ✅ Freeze top row
            sheet.freeze_panes = "A2"

            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = max(
                    (len(str(cell.value)) if cell.value else 0) for cell in column_cells
                )
                adjusted_width = min(30, max(12, max_length + 2))  # 👈 cap width at 30
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = adjusted_width

            # ✅ Style headers
            for cell in sheet[1]:
                cell.font = Font(bold=True, name="Calibri")
                cell.alignment = center
                cell.fill = header_fill
                cell.border = border

            # ✅ Style data rows
            for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                is_even_row = row[0].row % 2 == 0
                for idx, cell in enumerate(row):
                    cell.alignment = center
                    cell.border = border
                    if is_even_row:
                        cell.fill = alt_fill
                    # Make URL clickable (column 1)
                    if (
                        idx == 0
                        and isinstance(cell.value, str)
                        and cell.value.startswith("http")
                    ):
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0563C1", underline="single")

        wb.save(outfile)
        print(
            f"✅ Saved {sum(len(df) for df in self.sheets.values())} rows → {outfile}"
        )

    def extract_domain_from_url(self, product_dic: Dict[str, any]):
        return ((str(product_dic["url"])).split("/"))[2]

    def log_progress(self, product_dic: Dict[str, any], isLoggingInPerKg: False):

        # Domain like Basalam.com
        file_to_save_name = self.extract_domain_from_url(product_dic) + ".csv"

        # Create directory
        folder_path_to_save = "log"
        if isLoggingInPerKg:
            folder_path_to_save = os.path.join(folder_path_to_save, "per_kg")

        os.makedirs(folder_path_to_save, exist_ok=True)
        complete_path_to_file = os.path.join(folder_path_to_save, file_to_save_name)

        # get data frame to save the progress
        df = None
        try:
            if os.path.exists(complete_path_to_file):
                df = pd.read_csv(complete_path_to_file)
                df = pd.concat([df, pd.DataFrame([product_dic])])
            else:
                df = pd.DataFrame([product_dic])

            # Save the data frame as csv
            df.to_csv(path_or_buf=complete_path_to_file, index=False)

        except Exception as e:
            # If any error happens
            traceback.print_exc()
            error_path = os.path.join(folder_path_to_save, "error", file_to_save_name)

            # error data frame
            error_df = pd.DataFrame([product_dic])

            # concat it with past error data frame
            if os.path.exists(error_path):
                try:
                    exist_error_df = pd.read_csv(error_path)
                    error_df = pd.concat(error_df, exist_error_df)
                except Exception as e:
                    traceback.print_exc()
            error_df.to_csv(error_path, index=False)

        return

    def get_log_progress(self, product_dic: Dict[str, any], isLoggingInPerKg: False):

        # Domain like Basalam.com
        file_to_save_name = self.extract_domain_from_url(product_dic) + ".csv"
        folder_path_to_save = "log"
        if isLoggingInPerKg:
            folder_path_to_save = os.path.join(folder_path_to_save, "per_kg")

        complete_path_to_file = os.path.join(folder_path_to_save, file_to_save_name)

        # Create directory
        if os.path.exists(complete_path_to_file):
            return pd.read_csv(complete_path_to_file)

        else:
            return None

    async def isThisBlocksPageCheckedBefore(
        self, product_dic, isLoggingInPerKg: False
    ) -> bool:
        # Domain like Basalam.com
        isChecked = False
        file_complete_path = os.path.join(
            "log", self.extract_domain_from_url(product_dic) + ".csv"
        )
        if isLoggingInPerKg:
            file_complete_path = os.path.join(
                "log", "per_kg", self.extract_domain_from_url(product_dic) + ".csv"
            )

        try:
            if os.path.exists(file_complete_path):
                df = pd.read_csv(file_complete_path)

                # path exists and you got the data frame
                filtered_df = df[
                    df["product_menu_url"] == product_dic["product_menu_url"]
                ]

                if not filtered_df.empty:
                    # if any record is available with this menu url

                    if (
                        df.iloc[-1]["product_menu_url"]
                        != product_dic["product_menu_url"]
                        or "209"
                        == (str(df.iloc[-1]["product_menu_url"])).split("=")[-1].strip()
                    ):
                        # is exist in the list and isn't the last means that its task is completed
                        isChecked = True

        except Exception as e:
            traceback.print_exc()
            return False

        return isChecked
